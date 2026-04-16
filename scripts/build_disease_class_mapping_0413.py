import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "archive" / "processing_history" / "disease_update_new" / "disease_classify_all_update.xlsx"
INPUT_JSONL = ROOT / "data" / "processed" / "tekg2" / "tekg2_0413_clean.jsonl"
OUTPUT_DIR = ROOT / "data" / "processed" / "tekg2"
OUTPUT_JSON = OUTPUT_DIR / "disease_class_mapping_0413.json"
OUTPUT_TSV = OUTPUT_DIR / "disease_class_mapping_0413.tsv"
OUTPUT_REPORT = OUTPUT_DIR / "disease_class_mapping_0413_report.json"

GREEN_RGB = "FF92D050"
RED_RGB = "FFFF0000"


def normalize_name(value: str) -> str:
    return " ".join((value or "").strip().split()).casefold()


def row_category_path(row_values) -> list[str]:
    categories = []
    for value in row_values[1:9]:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            categories.append(text)
    return categories


def row_fill_rgb(cell) -> str | None:
    fill = getattr(cell, "fill", None)
    color = getattr(fill, "fgColor", None)
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb:
        return str(rgb)
    return None


def load_jsonl_disease_index(path: Path):
    by_normalized: dict[str, dict] = {}
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            raw = line.strip()
            if not raw:
                continue
            record = json.loads(raw)
            for disease in (record.get("entities") or {}).get("diseases", []) or []:
                name = (disease.get("name") or "").strip()
                if not name:
                    continue
                normalized = normalize_name(name)
                payload = by_normalized.setdefault(
                    normalized,
                    {
                        "jsonl_names": set(),
                        "description_candidates": set(),
                    },
                )
                payload["jsonl_names"].add(name)
                description = (disease.get("description") or "").strip()
                if description:
                    payload["description_candidates"].add(description)
    return by_normalized


def build_jsonl_payload(jsonl_entry: dict | None) -> dict:
    if not jsonl_entry:
        return {
            "matched_in_jsonl": False,
            "target_jsonl_name": None,
            "jsonl_names": [],
            "jsonl_description_candidates": [],
            "suggested_jsonl_description": None,
        }
    jsonl_names = sorted(jsonl_entry["jsonl_names"], key=lambda item: item.casefold())
    descriptions = sorted(jsonl_entry["description_candidates"], key=lambda item: (-len(item), item.casefold()))
    return {
        "matched_in_jsonl": True,
        "target_jsonl_name": jsonl_names[0] if jsonl_names else None,
        "jsonl_names": jsonl_names,
        "jsonl_description_candidates": descriptions,
        "suggested_jsonl_description": descriptions[0] if descriptions else None,
    }


def collect_sheet_rows(workbook_path: Path):
    workbook = load_workbook(workbook_path, data_only=True)

    found_rows = []
    found_sheet = workbook["Found"]
    for row in found_sheet.iter_rows(min_row=2):
        disease = row[0].value
        if not disease:
            continue
        found_rows.append(
            {
                "sheet": "Found",
                "disease": str(disease).strip(),
                "category_path": row_category_path([cell.value for cell in row]),
                "description": str(row[9].value).strip() if row[9].value else "",
            }
        )

    green_rows = []
    red_rows = []
    not_found_sheet = workbook["Not Found"]
    for row in not_found_sheet.iter_rows(min_row=2):
        disease = row[0].value
        if not disease:
            continue
        rgb = row_fill_rgb(row[0])
        payload = {
            "sheet": "Not Found",
            "disease": str(disease).strip(),
            "category_path": row_category_path([cell.value for cell in row]),
            "description": str(row[9].value).strip() if row[9].value else "",
            "fill_rgb": rgb,
        }
        if rgb == GREEN_RGB:
            green_rows.append(payload)
        elif rgb == RED_RGB:
            red_rows.append(payload)
    return found_rows, green_rows, red_rows


def aggregate_rows(rows: list[dict], mapping_type: str, jsonl_index: dict[str, dict]):
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[normalize_name(row["disease"])].append(row)

    entries = []
    for normalized, grouped_rows in sorted(grouped.items(), key=lambda item: item[0]):
        disease_name = grouped_rows[0]["disease"]
        category_paths = []
        seen_paths = set()
        descriptions = []
        for row in grouped_rows:
            path = tuple(row["category_path"])
            if path and path not in seen_paths:
                seen_paths.add(path)
                category_paths.append(list(path))
            description = row["description"]
            if description and description not in descriptions:
                descriptions.append(description)
        jsonl_payload = build_jsonl_payload(jsonl_index.get(normalized))
        entries.append(
            {
                "mapping_type": mapping_type,
                "disease_name": disease_name,
                "normalized_name": normalized,
                "category_path_count": len(category_paths),
                "category_paths": category_paths,
                "workbook_description_candidates": descriptions,
                "workbook_description": descriptions[0] if descriptions else None,
                "source_row_count": len(grouped_rows),
                **jsonl_payload,
            }
        )
    return entries


def build_red_notes(rows: list[dict], jsonl_index: dict[str, dict]):
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        key = " | ".join(row["category_path"])
        grouped[key].append(row)

    notes = []
    for category_path_key, grouped_rows in sorted(grouped.items(), key=lambda item: item[0].casefold()):
        workbook_names = [row["disease"] for row in grouped_rows]
        matched_jsonl_names = []
        suggested_canonical_name = None
        suggested_description = None
        for name in workbook_names:
            payload = jsonl_index.get(normalize_name(name))
            if payload:
                jsonl_names = sorted(payload["jsonl_names"], key=lambda item: item.casefold())
                for item in jsonl_names:
                    if item not in matched_jsonl_names:
                        matched_jsonl_names.append(item)
                if suggested_canonical_name is None and jsonl_names:
                    suggested_canonical_name = jsonl_names[0]
                if suggested_description is None and payload["description_candidates"]:
                    suggested_description = sorted(payload["description_candidates"], key=lambda item: (-len(item), item.casefold()))[0]
        notes.append(
            {
                "category_path": grouped_rows[0]["category_path"],
                "workbook_names": workbook_names,
                "matched_jsonl_names": matched_jsonl_names,
                "suggested_canonical_name": suggested_canonical_name,
                "suggested_jsonl_description": suggested_description,
                "workbook_description_candidates": [
                    row["description"] for row in grouped_rows if row["description"]
                ],
            }
        )
    return notes


def write_tsv(path: Path, rows: list[dict]):
    headers = [
        "mapping_type",
        "disease_name",
        "target_jsonl_name",
        "matched_in_jsonl",
        "category_path_count",
        "category_paths",
        "workbook_description",
        "suggested_jsonl_description",
        "source_row_count",
    ]
    lines = ["\t".join(headers)]
    for row in rows:
        values = [
            row["mapping_type"],
            row["disease_name"] or "",
            row["target_jsonl_name"] or "",
            "1" if row["matched_in_jsonl"] else "0",
            str(row["category_path_count"]),
            " || ".join(" > ".join(path) for path in row["category_paths"]),
            row["workbook_description"] or "",
            row["suggested_jsonl_description"] or "",
            str(row["source_row_count"]),
        ]
        lines.append("\t".join(values))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    jsonl_index = load_jsonl_disease_index(INPUT_JSONL)
    found_rows, green_rows, red_rows = collect_sheet_rows(WORKBOOK)

    found_entries = aggregate_rows(found_rows, "found_direct", jsonl_index)
    green_entries = aggregate_rows(green_rows, "green_multiclass", jsonl_index)
    red_notes = build_red_notes(red_rows, jsonl_index)

    matched_found = [entry for entry in found_entries if entry["matched_in_jsonl"]]
    unmatched_found = [entry for entry in found_entries if not entry["matched_in_jsonl"]]
    matched_green = [entry for entry in green_entries if entry["matched_in_jsonl"]]
    unmatched_green = [entry for entry in green_entries if not entry["matched_in_jsonl"]]

    mapping_rows = matched_found + matched_green
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_files": {
            "workbook": str(WORKBOOK),
            "jsonl": str(INPUT_JSONL),
        },
        "summary": {
            "found_total": len(found_entries),
            "found_matched_in_jsonl": len(matched_found),
            "found_unmatched_in_jsonl": len(unmatched_found),
            "green_total": len(green_entries),
            "green_matched_in_jsonl": len(matched_green),
            "green_unmatched_in_jsonl": len(unmatched_green),
            "red_duplicate_groups": len(red_notes),
            "mapping_rows_ready_for_step2": len(mapping_rows),
        },
        "found_direct_mappings": matched_found,
        "green_multiclass_mappings": matched_green,
        "found_unmatched": unmatched_found,
        "green_unmatched": unmatched_green,
        "red_duplicate_notes": red_notes,
    }

    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    OUTPUT_REPORT.write_text(json.dumps(payload["summary"], ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_tsv(OUTPUT_TSV, mapping_rows)
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
