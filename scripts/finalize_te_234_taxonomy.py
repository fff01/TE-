from __future__ import annotations

import csv
import json
import re
from copy import deepcopy
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

TREE_FILES = [
    ROOT / "transposon_tree" / "tree_all_4.27.txt",
    ROOT / "transposon_tree" / "tree_rmsk_repbase_4.27_2.txt",
]
WORKBOOK_PATH = ROOT / "transposon_tree" / "te_234_template2.xlsx"
SEED_JSON_PATH = ROOT / "data" / "processed" / "tekg2" / "tekg2_seed.json"
OUTPUT_CSV_PATH = ROOT / "data" / "statistics" / "te_234_template.csv"
OUTPUT_SUMMARY_PATH = ROOT / "data" / "statistics" / "te_234_cleanup_summary.json"
CONFIG_PATH = ROOT / "api" / "config.local.php"

TREE_TOKEN_RENAMES = {
    "HERV16": "HERV-16",
    "HERVI": "HERV-I",
    "FLAM_C": "FLAM C",
    "HURRS-P": "HuRRS-P",
    "L1M2A": "L1M2a",
    "L2B": "L2b",
    "AmnSINE1_HS": "Amnsine1_Hs",
    "asymmetric transposon end": "Asymmetric transposon end",
    "CHARLIE9": "Charlie9",
    "TIGGER1": "Tigger1",
    "TIGGER2": "Tigger2",
    "TIGGER3": "Tigger3",
    "TIGGER3a": "Tigger3a",
    "HSMAR1": "Hsmar1",
    "HSMAR2": "Hsmar2",
    "GOLEM": "Golem",
    "MADE1": "Made1",
    "LOOPER": "Looper",
}

TE_CANONICAL_MERGES = {
    "Transposable element": "DNA transposons",
    "Transposon": "DNA transposons",
    "HERV-L": "ERVL",
    "HERVL": "ERVL",
    "L1": "LINE-1",
    "retroelement": "retroposon",
}

MOVE_TO_GENE = {"HSATII"}
EXCLUDE_NON_HUMAN = {"L1Md-A5"}

DUPLICATE_ROW_PREFERENCES = {
    "LTR7A": "Family: HERVH",
    "LTR7B": "Family: HERVH",
    "LTR7C": "Family: HERVH",
    "LTR7Y": "Family: HERVH",
    "MER21C": "Family: MER21-group",
}

CSV_FIELDS = ["TE", "Class", "Subclass", "Order", "Superfamily", "Family", "Subfamily", "Subclade"]
PREFIX_TO_FIELD = {
    "Class": "Class",
    "Subclass": "Subclass",
    "Order": "Order",
    "Superfamily": "Superfamily",
    "Family": "Family",
    "Subfamily": "Subfamily",
    "Subclade": "Subclade",
}


def load_config() -> dict[str, str]:
    text = CONFIG_PATH.read_text(encoding="utf-8")
    config: dict[str, str] = {}
    for key in ("neo4j_url", "neo4j_user", "neo4j_password"):
        match = re.search(rf"'{key}'\s*=>\s*'([^']*)'", text)
        if match:
            config[key] = match.group(1)
    return config


def run_statement(config: dict[str, str], statement: str, parameters: dict[str, Any] | None = None) -> dict[str, Any]:
    payload = {"statements": [{"statement": statement, "parameters": parameters or {}}]}
    response = requests.post(
        config["neo4j_url"],
        auth=(config["neo4j_user"], config["neo4j_password"]),
        json=payload,
        timeout=180,
    )
    response.raise_for_status()
    data = response.json()
    if data.get("errors"):
        raise RuntimeError(json.dumps(data["errors"], ensure_ascii=False))
    return data


def patch_tree_file(path: Path, rename_map: dict[str, str]) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as fh:
        lines = fh.readlines()

    replacements: list[dict[str, str]] = []
    updated: list[str] = []
    for line in lines:
        line_ending = ""
        if line.endswith("\r\n"):
            body = line[:-2]
            line_ending = "\r\n"
        elif line.endswith("\n"):
            body = line[:-1]
            line_ending = "\n"
        else:
            body = line
        prefix_len = len(body) - len(body.lstrip(" │├└─"))
        prefix = body[:prefix_len]
        label = body[prefix_len:]
        new_label = label

        if ": " in label:
            label_prefix, label_value = label.split(": ", 1)
            if label_value in rename_map:
                new_label = f"{label_prefix}: {rename_map[label_value]}"
        elif label in rename_map:
            new_label = rename_map[label]

        if new_label != label:
            replacements.append({"from": label, "to": new_label})
        updated.append(prefix + new_label + line_ending)

    with path.open("w", encoding="utf-8", newline="") as fh:
        fh.writelines(updated)

    return replacements


def load_seed_payload() -> dict[str, Any]:
    return json.loads(SEED_JSON_PATH.read_text(encoding="utf-8"))


def load_template_rows() -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    ws = wb["te_234_template2"]
    rows: list[dict[str, Any]] = []
    by_name: dict[str, list[dict[str, Any]]] = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        entry = {
            "rownum": idx,
            "te_name": str(row[0]).strip(),
            "levels": [str(cell).strip() for cell in row[1:] if cell not in (None, "")],
        }
        rows.append(entry)
        by_name.setdefault(entry["te_name"], []).append(entry)
    return rows, by_name


def resolve_template_row(name: str, row_map: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    candidates = row_map.get(name, [])
    if not candidates:
        raise KeyError(f"Template row not found for {name}")
    if len(candidates) == 1:
        return candidates[0]
    preferred_marker = DUPLICATE_ROW_PREFERENCES.get(name)
    if preferred_marker:
        for candidate in candidates:
            if preferred_marker in candidate["levels"]:
                return candidate
    return candidates[0]


def row_to_csv_record(te_name: str, row: dict[str, Any]) -> dict[str, str]:
    record = {field: "" for field in CSV_FIELDS}
    record["TE"] = te_name
    for level in row["levels"]:
        if ": " not in level:
            continue
        prefix, value = level.split(": ", 1)
        normalized_prefix = prefix
        if prefix.startswith("Class "):
            normalized_prefix = "Class"
        elif prefix.startswith("Subclass "):
            normalized_prefix = "Subclass"
        field = PREFIX_TO_FIELD.get(normalized_prefix)
        if field and not record[field]:
            clean_value = value.strip()
            clean_value = TREE_TOKEN_RENAMES.get(clean_value, clean_value)
            record[field] = clean_value
    return record


def build_final_seed_and_csv(seed_payload: dict[str, Any], row_map: dict[str, list[dict[str, Any]]]) -> tuple[dict[str, Any], list[dict[str, str]], dict[str, Any]]:
    payload = deepcopy(seed_payload)
    transposons = payload["nodes"]["transposons"]
    genes = payload["nodes"]["genes"]

    gene_names = {gene["name"] if isinstance(gene, dict) else str(gene) for gene in genes}
    final_transposons: list[dict[str, Any]] = []
    final_csv_rows: list[dict[str, str]] = []
    seen_te: set[str] = set()

    merged_detected = []
    excluded_detected = []
    moved_to_gene = []
    duplicate_resolutions = {}

    for raw_item in transposons:
        item = deepcopy(raw_item)
        raw_name = item.get("name") if isinstance(item, dict) else str(item)

        if raw_name in MOVE_TO_GENE:
            excluded_detected.append({"name": raw_name, "action": "move_to_gene"})
            moved_to_gene.append(raw_name)
            if raw_name not in gene_names:
                genes.append(
                    {
                        "name": raw_name,
                        "description": item.get("description", "") if isinstance(item, dict) else "",
                    }
                )
                gene_names.add(raw_name)
            continue

        if raw_name in EXCLUDE_NON_HUMAN:
            excluded_detected.append({"name": raw_name, "action": "exclude_non_human"})
            continue

        canonical = TE_CANONICAL_MERGES.get(raw_name, raw_name)
        if canonical != raw_name:
            merged_detected.append({"from": raw_name, "to": canonical})

        if canonical in seen_te:
            continue

        template_row = resolve_template_row(canonical, row_map)
        if len(row_map.get(canonical, [])) > 1:
            duplicate_resolutions[canonical] = template_row["rownum"]

        if canonical != raw_name and isinstance(item, dict):
            item["name"] = canonical

        if canonical == raw_name:
            # Prefer the canonical entity's own description if it appears later in the list.
            canonical_item = next(
                (
                    deepcopy(candidate)
                    for candidate in transposons
                    if isinstance(candidate, dict) and candidate.get("name") == canonical
                ),
                None,
            )
            if canonical_item is not None:
                item = canonical_item
        elif isinstance(item, dict):
            canonical_item = next(
                (
                    candidate
                    for candidate in transposons
                    if isinstance(candidate, dict) and candidate.get("name") == canonical
                ),
                None,
            )
            if canonical_item is not None and canonical_item.get("description"):
                item["description"] = canonical_item["description"]

        final_transposons.append(item)
        final_csv_rows.append(row_to_csv_record(canonical, template_row))
        seen_te.add(canonical)

    payload["nodes"]["transposons"] = final_transposons

    summary = {
        "final_te_count": len(final_transposons),
        "configured_merges": [{"from": key, "to": value} for key, value in TE_CANONICAL_MERGES.items()],
        "configured_exclusions": {
            "move_to_gene": sorted(MOVE_TO_GENE),
            "exclude_non_human": sorted(EXCLUDE_NON_HUMAN),
        },
        "detected_merges_in_seed": merged_detected,
        "detected_exclusions_in_seed": excluded_detected,
        "moved_to_gene": moved_to_gene,
        "duplicate_resolutions": duplicate_resolutions,
    }
    return payload, final_csv_rows, summary


def write_seed_payload(payload: dict[str, Any]) -> None:
    SEED_JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(rows: list[dict[str, str]]) -> None:
    with OUTPUT_CSV_PATH.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def merge_te_name(config: dict[str, str], drop: str, canonical: str) -> None:
    statement = """
MERGE (keep:TE {name: $canonical})
WITH keep
MATCH (drop:TE {name: $drop})
WITH keep, drop
WHERE elementId(keep) <> elementId(drop)
CALL {
  WITH keep, drop
  OPTIONAL MATCH (src)-[r:SUBFAMILY_OF]->(drop)
  FOREACH (_ IN CASE WHEN r IS NULL THEN [] ELSE [1] END |
    MERGE (src)-[:SUBFAMILY_OF]->(keep)
    DELETE r
  )
}
CALL {
  WITH keep, drop
  OPTIONAL MATCH (drop)-[r:SUBFAMILY_OF]->(dst)
  FOREACH (_ IN CASE WHEN r IS NULL THEN [] ELSE [1] END |
    MERGE (keep)-[:SUBFAMILY_OF]->(dst)
    DELETE r
  )
}
CALL {
  WITH keep, drop
  OPTIONAL MATCH (src)-[r:BIO_RELATION]->(drop)
  FOREACH (_ IN CASE WHEN r IS NULL THEN [] ELSE [1] END |
    MERGE (src)-[r2:BIO_RELATION {predicate: r.predicate}]->(keep)
    SET r2.pmids = reduce(acc = coalesce(r2.pmids, []), x IN coalesce(r.pmids, []) | CASE WHEN x IN acc THEN acc ELSE acc + x END),
        r2.source_group = coalesce(r2.source_group, r.source_group)
    DELETE r
  )
}
CALL {
  WITH keep, drop
  OPTIONAL MATCH (drop)-[r:BIO_RELATION]->(dst)
  FOREACH (_ IN CASE WHEN r IS NULL THEN [] ELSE [1] END |
    MERGE (keep)-[r2:BIO_RELATION {predicate: r.predicate}]->(dst)
    SET r2.pmids = reduce(acc = coalesce(r2.pmids, []), x IN coalesce(r.pmids, []) | CASE WHEN x IN acc THEN acc ELSE acc + x END),
        r2.source_group = coalesce(r2.source_group, r.source_group)
    DELETE r
  )
}
SET keep.description = CASE WHEN coalesce(keep.description, '') = '' THEN drop.description ELSE keep.description END
DETACH DELETE drop
"""
    run_statement(config, statement, {"canonical": canonical, "drop": drop})


def relabel_hsatii_to_gene(config: dict[str, str]) -> None:
    statement = """
MATCH (n:TE {name: 'HSATII'})
REMOVE n:TE
SET n:Gene,
    n.source_group = 'genes',
    n.reclassified_reason = 'satellite_not_transposon'
"""
    run_statement(config, statement)


def relabel_l1md_a5_nonhuman(config: dict[str, str]) -> None:
    statement = """
MATCH (n:TE {name: 'L1Md-A5'})
REMOVE n:TE
SET n:NonHumanTE,
    n.source_group = 'nonhuman_transposons',
    n.reclassified_reason = 'non_human_transposon'
"""
    run_statement(config, statement)


def write_summary(summary: dict[str, Any]) -> None:
    OUTPUT_SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    tree_changes = {}
    for tree_path in TREE_FILES:
        tree_changes[str(tree_path.relative_to(ROOT))] = patch_tree_file(tree_path, TREE_TOKEN_RENAMES)

    _, row_map = load_template_rows()
    seed_payload = load_seed_payload()
    updated_seed, csv_rows, summary = build_final_seed_and_csv(seed_payload, row_map)
    write_seed_payload(updated_seed)
    write_csv(csv_rows)

    config = load_config()
    for drop, canonical in TE_CANONICAL_MERGES.items():
        merge_te_name(config, drop, canonical)
    relabel_hsatii_to_gene(config)
    relabel_l1md_a5_nonhuman(config)

    summary["tree_changes"] = tree_changes
    write_summary(summary)

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
