from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SEED_JSON = ROOT / "data" / "processed" / "tekg2" / "tekg2_seed.json"
XLSX_ALL = ROOT / "data" / "statistics" / "TE_tree_output_all.xlsx"
XLSX_BASE = ROOT / "data" / "statistics" / "TE_tree_output.xlsx"
OUT_CSV = ROOT / "data" / "statistics" / "te_234_classification.csv"


OUTPUT_FIELDS = [
    "te_name",
    "match_status",
    "source_workbook",
    "class",
    "subclass",
    "order",
    "superfamily",
    "family",
    "subfamily",
    "subclade",
    "category_1",
    "category_2",
    "root_1",
    "root_2",
]


def load_seed_te_names() -> list[str]:
    payload = json.loads(SEED_JSON.read_text(encoding="utf-8"))
    items = payload["nodes"]["transposons"]
    names: list[str] = []
    for item in items:
        if isinstance(item, dict):
            name = item.get("name") or item.get("label") or item.get("id") or ""
            names.append(str(name).strip())
        else:
            names.append(str(item).strip())
    return names


def parse_leaf_workbook(path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Leaf Paths"]
    rows: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        element = str(row[0]).strip()
        parsed = {
            "class": "",
            "subclass": "",
            "order": "",
            "superfamily": "",
            "family": "",
            "subfamily": "",
            "subclade": "",
            "category_1": "",
            "category_2": "",
            "root_1": "",
            "root_2": "",
        }

        generic_levels: list[str] = []
        roots: list[str] = []

        for cell in row[1:]:
            if not cell:
                continue
            value = str(cell).strip()
            if not value:
                continue
            if ": " in value:
                prefix, content = value.split(": ", 1)
                raw_prefix = prefix.strip().lower()
                key = raw_prefix.replace(" ", "_")
                if raw_prefix.startswith("class "):
                    key = "class"
                elif raw_prefix.startswith("subclass "):
                    key = "subclass"
                if key in parsed:
                    parsed[key] = content.strip()
                else:
                    generic_levels.append(value)
            else:
                lower_value = value.lower()
                if "transposable elements" in lower_value or "mobile genetic element" in lower_value:
                    roots.append(value)
                else:
                    generic_levels.append(value)

        if generic_levels:
            parsed["category_1"] = generic_levels[0]
        if len(generic_levels) > 1:
            parsed["category_2"] = generic_levels[1]
        if roots:
            parsed["root_1"] = roots[0]
        if len(roots) > 1:
            parsed["root_2"] = roots[1]

        rows[element] = parsed
    return rows


def main() -> None:
    seed_names = load_seed_te_names()
    all_rows = parse_leaf_workbook(XLSX_ALL)
    base_rows = parse_leaf_workbook(XLSX_BASE)

    output_rows: list[dict[str, str]] = []

    for name in seed_names:
        if name in all_rows:
            row = {
                "te_name": name,
                "match_status": "exact_match",
                "source_workbook": XLSX_ALL.name,
                **all_rows[name],
            }
        elif name in base_rows:
            row = {
                "te_name": name,
                "match_status": "exact_match",
                "source_workbook": XLSX_BASE.name,
                **base_rows[name],
            }
        else:
            row = {
                "te_name": name,
                "match_status": "no_exact_match",
                "source_workbook": "",
                "class": "",
                "subclass": "",
                "order": "",
                "superfamily": "",
                "family": "",
                "subfamily": "",
                "subclade": "",
                "category_1": "",
                "category_2": "",
                "root_1": "",
                "root_2": "",
            }
        output_rows.append(row)

    with OUT_CSV.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(output_rows)

    exact_matches = sum(1 for row in output_rows if row["match_status"] == "exact_match")
    print(f"Wrote: {OUT_CSV}")
    print(f"Rows: {len(output_rows)}")
    print(f"Exact matches: {exact_matches}")
    print(f"Unmatched: {len(output_rows) - exact_matches}")


if __name__ == "__main__":
    main()
