import json
from pathlib import Path
from collections import OrderedDict
import openpyxl

def load_standardization_map(excel_path, sheet_name="Sheet1"):
    """
    从 Excel 加载标准化映射。
    每行第一列为标准名称，后续非空单元格为同义词。
    返回字典 {同义词: 标准名称}，包含标准名称自身。
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    mapping = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        standard = str(row[0]).strip()
        if not standard:
            continue
        mapping[standard] = standard  # 标准名自身
        for cell in row[1:]:
            if cell is None:
                continue
            syn = str(cell).strip()
            if syn:
                mapping[syn] = standard
    return mapping

def normalize_string(s, mapping):
    """如果字符串在映射中则返回标准化名称，否则原样返回。"""
    if not isinstance(s, str):
        return s
    stripped = s.strip()
    return mapping.get(stripped, s)

def process_jsonl(input_path, output_jsonl_path, mapping):
    """
    处理 JSONL 文件：
    - 标准化疾病实体的 name
    - 标准化 relations 中的 source 和 target
    - 写入新的 JSONL 文件
    - 返回标准化后的疾病实体列表 [(name, description)] 去重
    """
    disease_dict = OrderedDict()  # name -> description
    total_lines = 0
    modified_lines = 0

    with open(input_path, 'r', encoding='utf-8') as infile, \
         open(output_jsonl_path, 'w', encoding='utf-8') as outfile:

        for line_num, line in enumerate(infile, 1):
            line = line.strip()
            if not line:
                continue
            total_lines += 1
            try:
                data = json.loads(line)
            except json.JSONDecodeError as e:
                print(f"警告：第 {line_num} 行 JSON 解析失败，已跳过。错误：{e}")
                continue

            modified = False

            # 1) 处理 entities.diseases
            entities = data.get('entities')
            if isinstance(entities, dict):
                diseases = entities.get('diseases')
                if isinstance(diseases, list):
                    for disease in diseases:
                        if not isinstance(disease, dict):
                            continue
                        orig_name = disease.get('name')
                        if orig_name:
                            new_name = normalize_string(orig_name, mapping)
                            if new_name != orig_name:
                                disease['name'] = new_name
                                modified = True
                            # 收集疾病实体（标准化后的名称）
                            std_name = new_name
                            desc = disease.get('description', '')
                            if std_name not in disease_dict:
                                disease_dict[std_name] = desc

            # 2) 处理 relations 中的 source 和 target
            relations = data.get('relations')
            if isinstance(relations, list):
                for rel in relations:
                    if not isinstance(rel, dict):
                        continue
                    # source
                    if 'source' in rel:
                        src = rel['source']
                        if isinstance(src, str):
                            new_src = normalize_string(src, mapping)
                            if new_src != src:
                                rel['source'] = new_src
                                modified = True
                    # target
                    if 'target' in rel:
                        tgt = rel['target']
                        if isinstance(tgt, str):
                            new_tgt = normalize_string(tgt, mapping)
                            if new_tgt != tgt:
                                rel['target'] = new_tgt
                                modified = True

            if modified:
                modified_lines += 1

            # 写入修改后的行
            outfile.write(json.dumps(data, ensure_ascii=False) + '\n')

    print(f"处理完成：总行数 {total_lines}，修改行数 {modified_lines}")
    # 将疾病字典转换为列表
    return [(name, desc) for name, desc in disease_dict.items()]

def write_excel(diseases_data, output_excel_path):
    """将疾病列表写入 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diseases"
    ws.append(["Disease Name", "Description"])
    for name, desc in diseases_data:
        ws.append([name, desc])
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60
    wb.save(output_excel_path)
    print(f"疾病实体 Excel 已保存：{output_excel_path}，共 {len(diseases_data)} 条")

def main():
    # 文件路径
    base_dir = Path(r"C:\Users\fongi\Desktop\TE")
    jsonl_path = base_dir / "data_update" / "te_kg2_final.jsonl"
    mapping_excel = base_dir / "disease_update" / "diseases.xlsx"
    output_jsonl = jsonl_path.parent / f"{jsonl_path.stem}_standardized.jsonl"
    output_excel = jsonl_path.parent / f"{jsonl_path.stem}_diseases_standardized.xlsx"

    if not jsonl_path.is_file():
        print(f"错误：找不到 JSONL 文件 {jsonl_path}")
        return
    if not mapping_excel.is_file():
        print(f"错误：找不到映射文件 {mapping_excel}")
        return

    print("加载标准化映射...")
    mapping = load_standardization_map(mapping_excel, sheet_name="Sheet1")
    print(f"映射表包含 {len(mapping)} 个词条（含同义词）")

    print("处理 JSONL 并生成标准化文件...")
    diseases_data = process_jsonl(jsonl_path, output_jsonl, mapping)

    if not diseases_data:
        print("警告：未找到任何疾病实体。")
    else:
        write_excel(diseases_data, output_excel)
        print(f"标准化后的 JSONL 已保存：{output_jsonl}")

if __name__ == "__main__":
    main()