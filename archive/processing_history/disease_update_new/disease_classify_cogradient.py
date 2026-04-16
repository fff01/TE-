import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
MANUAL_DROP_DIR = ROOT / "data" / "raw" / "manual_drop"
DISEASE_UPDATE_DIR = ROOT / "archive" / "processing_history" / "disease_update_new"

def copy_cell_style(source_cell, target_cell):
    """复制源单元格的样式到目标单元格（包括字体、填充、边框等）"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.alignment = copy(source_cell.alignment)

def copy_worksheet_to_workbook(source_ws, target_wb, target_sheet_name):
    """
    将源工作表完整复制到目标工作簿（包括数据、样式、列宽、行高）
    如果目标工作簿中已存在同名工作表，则先删除再创建。
    """
    # 如果目标工作簿中已有同名工作表，删除
    if target_sheet_name in target_wb.sheetnames:
        target_wb.remove(target_wb[target_sheet_name])
    # 新建工作表（临时名称，避免冲突）
    new_ws = target_wb.create_sheet(title="_temp_copy")
    # 复制列宽
    for col in source_ws.column_dimensions:
        new_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width
    # 复制行高
    for row in source_ws.row_dimensions:
        new_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height
    # 复制合并单元格
    for merged_range in source_ws.merged_cells.ranges:
        new_ws.merge_cells(str(merged_range))
    # 复制所有单元格的值和样式
    for row in range(1, source_ws.max_row + 1):
        for col in range(1, source_ws.max_column + 1):
            src_cell = source_ws.cell(row, col)
            tgt_cell = new_ws.cell(row, col)
            tgt_cell.value = src_cell.value
            copy_cell_style(src_cell, tgt_cell)
    # 重命名工作表为最终名称
    new_ws.title = target_sheet_name
    return new_ws

def main():
    src_file = str(MANUAL_DROP_DIR / "disease_classify_all.xlsx")
    tgt_file = str(DISEASE_UPDATE_DIR / "disease_classify_all_update.xlsx")
    sheet_name = "Not Found"
    color_sheet_name = "color"

    # 加载工作簿
    wb_src = openpyxl.load_workbook(src_file)
    wb_tgt = openpyxl.load_workbook(tgt_file)

    # 复制 color 工作表
    if color_sheet_name in wb_src.sheetnames:
        src_color_ws = wb_src[color_sheet_name]
        copy_worksheet_to_workbook(src_color_ws, wb_tgt, color_sheet_name)
        print(f"已复制工作表 '{color_sheet_name}' 到目标文件")
    else:
        print(f"源文件中没有找到工作表 '{color_sheet_name}'")

    # 处理 Not Found 工作表
    if sheet_name not in wb_src.sheetnames:
        print(f"源文件中找不到工作表 '{sheet_name}'")
        wb_tgt.save(tgt_file)
        return
    if sheet_name not in wb_tgt.sheetnames:
        print(f"目标文件中找不到工作表 '{sheet_name}'")
        wb_tgt.save(tgt_file)
        return

    ws_src = wb_src[sheet_name]
    ws_tgt = wb_tgt[sheet_name]

    # 获取源表标题行，确定 Description 列位置
    src_header = [cell.value for cell in ws_src[1]]
    try:
        src_desc_col = src_header.index("Description") + 1
    except ValueError:
        print("源表中找不到 'Description' 列")
        wb_tgt.save(tgt_file)
        return
    src_cat_start = 2          # B列
    src_cat_end = src_desc_col - 1

    # 构建源数据字典：小写疾病名 -> (原始疾病名, 行号, 分类单元格列表)
    src_data = {}
    for row in range(2, ws_src.max_row + 1):
        disease_cell = ws_src.cell(row, 1)
        disease = disease_cell.value
        if not disease:
            continue
        cat_cells = [ws_src.cell(row, col) for col in range(src_cat_start, src_cat_end + 1)]
        src_data[disease.lower()] = (disease, row, cat_cells)

    # 获取目标表结构
    tgt_header = [cell.value for cell in ws_tgt[1]]
    try:
        tgt_desc_col = tgt_header.index("Description") + 1
    except ValueError:
        print("目标表中找不到 'Description' 列")
        wb_tgt.save(tgt_file)
        return
    tgt_cat_start = 2
    tgt_cat_end = tgt_desc_col - 1
    current_cat_count = tgt_cat_end - tgt_cat_start + 1

    # 遍历目标表，更新匹配的行
    for row in range(2, ws_tgt.max_row + 1):
        disease_cell = ws_tgt.cell(row, 1)
        disease = disease_cell.value
        if not disease:
            continue
        disease_lower = disease.lower()
        if disease_lower not in src_data:
            continue

        src_disease_name, src_row, src_cat_cells = src_data[disease_lower]
        src_cat_count = len(src_cat_cells)

        # 如果源分类列数多于当前目标分类列数，插入新列
        if src_cat_count > current_cat_count:
            need_insert = src_cat_count - current_cat_count
            for _ in range(need_insert):
                ws_tgt.insert_cols(tgt_desc_col)  # 在 Description 左侧插入一列
            tgt_cat_end = tgt_desc_col - 1
            current_cat_count = src_cat_count
            # 更新标题行，为新列添加空标题
            for i in range(1, need_insert + 1):
                new_col_idx = tgt_cat_start + current_cat_count - need_insert + i - 1
                new_header_cell = ws_tgt.cell(1, new_col_idx)
                new_header_cell.value = f"Empty_{current_cat_count - need_insert + i}"
                # 复制原有分类标题的字体样式（简单加粗）
                src_header_cell = ws_tgt.cell(1, tgt_cat_start)
                if src_header_cell.has_style:
                    new_header_cell.font = copy(src_header_cell.font)

        # 复制分类单元格的值和样式
        for col_offset, src_cell in enumerate(src_cat_cells):
            target_col = tgt_cat_start + col_offset
            target_cell = ws_tgt.cell(row, target_col)
            target_cell.value = src_cell.value
            copy_cell_style(src_cell, target_cell)

        # 可选：复制疾病名称列的样式（保持原样式一致性）
        src_disease_cell = ws_src.cell(src_row, 1)
        copy_cell_style(src_disease_cell, disease_cell)

    # 保存目标文件
    wb_tgt.save(tgt_file)
    print(f"处理完成，已更新 {tgt_file} 中的 '{sheet_name}' 工作表。")

if __name__ == "__main__":
    main()